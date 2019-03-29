from flask import Flask
from flask_restful import Resource, Api, reqparse
from openpyxl import load_workbook

app = Flask(__name__)
api = Api(app)

_row = 0
_column = 0

baseFolderAddr = "/home/gaon/excel"
baseFolderAddrCctv = "/home/gaon/cctv"


class dick(Resource):
	def get(self):
		global _row
		global _column
		global baseFolderAddr
		datas = []

		fileAddr = baseFolderAddr + '/' + str(_row) +str(_column)+".xlsx"
		try:
			locationEx = load_workbook(filename=fileAddr)
		except FileNotFoundError:
			return None
		Sheet1 = locationEx['sheet1']

		for i in Sheet1.rows:
			_longitude = i[2].value
			_latitude = i[3].value
			data = {"latitude":_latitude,"longitude":_longitude}
			datas.append(data)
		return datas

		# for k in range(_row - 1, _row + 2):
		# 	for j in range(_column - 1, _column + 2):
		# 		fileAddr = baseFolderAddr + '/' + str(k) + str(j) + ".xlsx"
		# 		try:
		# 			locationEx = load_workbook(filename=fileAddr)
		# 		except FileNotFoundError:
		# 			continue
		#
		# 		Sheet1 = locationEx['sheet1']
		#
		# 		for i in Sheet1.rows:
		# 			_longitude = i[2].value
		# 			_latitude = i[3].value
		# 			data = {"latitude": _latitude, "longitude": _longitude}
		# 			datas.append(data)
		#
		# return datas



	def post(self):

		global _row
		global _column

		parser = reqparse.RequestParser()
		parser.add_argument('row',type=int)
		parser.add_argument('column',type=int)
		args = parser.parse_args()
		_row = args['row']
		_column = args['column']

		return {'row':_row, 'column':_column}




class cctv(Resource):
	def get(self):
		global _row
		global _column
		global baseFolderAddrCctv
		datas=[]

		# fileAddr = baseFolderAddrCctv + '/' + str(_row) +str(_column)+".xlsx"
		# try:
		# 	locationEx = load_workbook(filename=fileAddr)
		# except FileNotFoundError:
		# 	return None
		# Sheet1 = locationEx['sheet1']
		#
		# for i in Sheet1.rows:
		# 	_longitude = i[2].value
		# 	_latitude = i[3].value
		# 	data = {"latitude":_latitude,"longitude":_longitude}
		# 	datas.append(data)
		# return datas


		for k in range(_row-1,_row+2):
			for j in range(_column-1,_column+2):
				fileAddr = baseFolderAddrCctv + '/' + str(k) + str(j) + ".xlsx"
				try:
					locationEx = load_workbook(filename=fileAddr)
				except FileNotFoundError:
					continue

				Sheet1 = locationEx['sheet1']


				for i in Sheet1.rows:
					_longitude = i[2].value
					_latitude = i[3].value
					data = {"latitude": _latitude, "longitude": _longitude}
					datas.append(data)


		return datas


	def post(self):

		global _row
		global _column

		parser = reqparse.RequestParser()
		parser.add_argument('row', type=int)
		parser.add_argument('column', type=int)
		args = parser.parse_args()
		_row = args['row']
		_column = args['column']

		return {'row': _row, 'column': _column}



api.add_resource(dick, '/lamp')
api.add_resource(cctv, '/cctv')


if __name__ == '__main__':
	app.run(debug=True)
